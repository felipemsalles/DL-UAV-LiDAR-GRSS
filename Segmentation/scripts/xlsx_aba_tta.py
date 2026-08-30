"""Write the test-time-adaptation tab into the detection workbook.

Reads what `exp_tta_comparison.py` recorded and builds a single tab, with the two
metrics side by side for both models. The whole tab is rewritten on every execution,
so running it twice duplicates nothing.

Usage:
    PYTHONPATH=. python scripts/xlsx_aba_tta.py [path/to/workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from greenvista import config

ABA = "Test-time adaptation"
PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"
VERDE = PatternFill("solid", fgColor="1D6B45")
CLARO = PatternFill("solid", fgColor="EDF3EF")
BRANCO = Font(bold=True, color="FFFFFF", sz=10)
LIMIAR_TABELA = 2.0     # the threshold the workbook shows; the CSV keeps the full sweep

NOMES = {"baseline": "no adaptation",
         "baseline_julho": "no adaptation, July run",
         "baseline_pareado": "no adaptation, paired run",
         "adabn": "with AdaBN",
         "tent": "with TENT"}
ORDEM = ["baseline", "baseline_pareado", "baseline_julho", "adabn", "tent"]


def chave(cond):
    """Order as no-adaptation, then AdaBN, then TENT, rather than alphabetically."""
    return (ORDEM.index(cond) if cond in ORDEM else len(ORDEM), cond)


def cabecalho(ws, linha, colunas):
    for j, t in enumerate(colunas, 1):
        c = ws.cell(linha, j, t)
        c.fill, c.font = VERDE, BRANCO
        c.alignment = Alignment(wrap_text=True, vertical="center")


def corpo(ws, linha, valores, formatos=None):
    for j, v in enumerate(valores, 1):
        fmt = formatos[j - 1] if formatos else None
        # a hit rate of exactly 100% arrives as 1.0 and openpyxl writes it as an integer,
        # making the cell show "1" instead of "100.0%"; forcing float keeps the percentage.
        if fmt and fmt.endswith("%"):
            v = float(v)
        c = ws.cell(linha, j, v)
        c.fill = CLARO
        if fmt:
            c.number_format = fmt


def titulo(ws, linha, texto, sz=11):
    c = ws.cell(linha, 1, texto)
    c.font = Font(bold=True, sz=sz)
    return linha + 1


def bloco_casada(ws, ln, tot, titulo_txt, explicacao):
    ln = titulo(ws, ln, titulo_txt)
    ln = titulo(ws, ln, explicacao)
    cabecalho(ws, ln, ["Model", "Condition", "Stems", "Predictions", "Hits",
                       "False positive", "Missed", "Precision", "Recall", "F1"])
    ln += 1
    sub = tot[tot.limiar_m == LIMIAR_TABELA].copy()
    sub["_o"] = sub.condicao.map(chave)
    for _, r in sub.sort_values(["modelo", "_o"]).iterrows():
        corpo(ws, ln, [r.modelo, NOMES.get(r.condicao, r.condicao), int(r.ref),
                       int(r.pred), int(r.TP), int(r.FP), int(r.FN),
                       float(r.precisao), float(r.revocacao), float(r.F1)],
              [None] * 7 + ["0.0%", "0.0%", "0.0%"])
        ln += 1
    return ln + 1


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    livro = Path(args[0]) if args else PADRAO
    cont = pd.read_csv(config.OUT_DIR / "tta_comparison_contagem.csv")
    tot = pd.read_csv(config.OUT_DIR / "tta_comparison_casada_total.csv")
    base_tot = config.OUT_DIR / "tta_base_casada_total.csv"
    base_tot = pd.read_csv(base_tot) if base_tot.exists() else None

    wb = openpyxl.load_workbook(livro)
    if ABA in wb.sheetnames:
        del wb[ABA]
    ws = wb.create_sheet(ABA)
    for col, larg in zip("ABCDEFGHI", [30, 13, 11, 11, 11, 11, 11, 11, 11]):
        ws.column_dimensions[col].width = larg

    ln = titulo(ws, 1, "Test-time adaptation, the two segmenters", sz=14)
    ln = titulo(ws, ln, "Stand 001, the same 18 views of 32 m for all of them. "
                        "Adapting at test time means letting the model recalibrate "
                        "on the cloud it is seeing right now, with no labels at all.")
    ln += 1

    # ------------------------------------------------ count
    ln = titulo(ws, ln, "Count inside the plot, against the field census")
    ln = titulo(ws, ln, "Two rulers. The 12 m circle covers 452 m². The 11.28 m one "
                        "closes exactly the "
                        "400 m² the field crew measured, and that is why it is the fair number.")
    cabecalho(ws, ln, ["Model", "Condition", "Plot", "Field",
                       "Found r 12 m", "Hit rate r 12 m",
                       "Found r 11.28 m", "Hit rate r 11.28 m"])
    ln += 1
    cont = cont.copy()
    cont["_o"] = cont.condicao.map(chave)
    for _, r in cont.sort_values(["modelo", "_o", "parcela"]).iterrows():
        corpo(ws, ln, [r.modelo, NOMES.get(r.condicao, r.condicao), r.parcela,
                       int(r.campo), int(r.n_r12), float(r.det_r12),
                       int(r.n_r400), float(r.det_r400)],
              [None, None, None, None, None, "0.0%", None, "0.0%"])
        ln += 1
    ln += 1

    # ------------------------------------------------ matched metric
    ln = bloco_casada(
        ws, ln, tot,
        f"Matched metric, threshold of {LIMIAR_TABELA:.0f} m, the two plots pooled",
        "Each prediction is matched with at most one stem of the TLS map, by "
        "optimal assignment. Here what was left out can be separated from what was "
        "invented, something the count alone hides. The predicted position is the "
        "centroid of the instance.")

    if base_tot is not None:
        ln = bloco_casada(
            ws, ln, base_tot,
            "The same metric, with the position taken at the stem base",
            "The crown centroid sits metres away from the trunk when the tree is "
            "leaning or the crown is asymmetric, and the reference is of stems. Here "
            "the position comes from the mean of the points in the first three metres above the "
            "ground, the same rule in both models. Compare with the block above: what "
            "rises here was misalignment, not a missed tree.")

    ws.freeze_panes = "A2"
    wb.save(livro)
    print(f"tab '{ABA}' rewritten in {livro}")


if __name__ == "__main__":
    main()
