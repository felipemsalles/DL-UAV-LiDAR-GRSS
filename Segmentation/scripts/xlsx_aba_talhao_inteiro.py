"""Write the "Whole stand" tab from the wall-to-wall run.

Carries the stem count over the 9.87 ha covered by the point cloud, not only over the
13 plots (717 census stems in 0.52 ha).

Two caveats are printed on the tab itself, next to the number:

  1. Outside the plots there is no field check. The stand total is a model measurement,
     not a verified one. The verifiable part is the one falling inside the plots, in the
     comparison column.
  2. The cloud was clipped exactly at the stand boundary, with no margin, measured point
     by point. Trees on the outer row have their crown truncated for lack of data. The
     3 m ring column indicates what fraction of the count is in that condition: 12.9%,
     against 12.5% of the area.

Usage: PYTHONPATH=. python scripts/xlsx_aba_talhao_inteiro.py [workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from greenvista import config

ABA = "Whole stand"
PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"
FONTE = config.OUT_DIR / "w2w_por_talhao.csv"
VERDE = PatternFill("solid", fgColor="1D6B45")
CLARO = PatternFill("solid", fgColor="EDF3EF")
BRANCO = Font(bold=True, color="FFFFFF", sz=10)
NOME = {"001": "001", "002": "002", "004": "004 (young)", "005": "005",
        "006": "006", "007": "007", "TOTAL": "Total"}

COLS = [("Stand", 13), ("Area (ha)", 10), ("Trees", 10), ("Trees per ha", 15),
        ("Boundary (m/ha)", 14), ("In the 3 m ring", 15), ("% in the ring", 10),
        ("Views per tree", 17)]


def main():
    livro = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    if not FONTE.exists():
        raise SystemExit(f"could not find {FONTE}. Run exp_w2w_contagem.py first")
    d = pd.read_csv(FONTE)

    wb = openpyxl.load_workbook(livro)
    if ABA in wb.sheetnames:
        del wb[ABA]
    ws = wb.create_sheet(ABA, 2)          # right after the summary
    for (t, w), col in zip(COLS, "ABCDEFGH"):
        ws.column_dimensions[col].width = w

    ws.cell(1, 1, "Trees in the whole stand").font = Font(bold=True, sz=14)
    ws.cell(2, 1, "SegmentAnyTree with adaptation, 9 views, sweeping the whole stand and not "
                  "only the surroundings of the plots. Fusion at a 1.7 m radius, chosen against the "
                  "stem map.").font = Font(sz=10, italic=True)

    ln = 4
    for j, (t, _) in enumerate(COLS, 1):
        c = ws.cell(ln, j, t)
        c.fill, c.font = VERDE, BRANCO
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[ln].height = 30
    ln += 1

    for _, r in d.iterrows():
        total = str(r.talhao) == "TOTAL"
        vals = [NOME.get(str(r.talhao), str(r.talhao)), float(r.ha), int(r.arvores),
                float(r.por_ha), int(r.m_divisa_ha), int(r.no_anel_3m),
                float(r.pct_anel) / 100, float(r.vistas_por_arvore)]
        for j, v in enumerate(vals, 1):
            c = ws.cell(ln, j, v)
            c.fill = CLARO
            if j == 7:
                c.number_format = "0.0%"
            elif j in (2, 4, 8):
                c.number_format = "0.0"
            if total:
                c.font = Font(bold=True)
        ln += 1

    ln += 1
    for txt in [
        "How to read",
        "Outside the plots there is no field count, so the total of each stand is a "
        "model measurement and not a checked number. What can be checked is the part falling "
        "inside the 13 plots, and that comparison is in the summary tab.",
        "The 3 m ring column separates the trees of the outer strip of the stand. The delivered cloud "
        "was clipped exactly at the boundary, with no metre of margin, so the crown of those trees "
        "is truncated for lack of data and not by a clipping choice. It is the part of the total that "
        "deserves the most suspicion, and the 3 m ring is 12.5% of the area.",
        "Boundary in metres per hectare measures edge exposure, and it is not the same across stands. 006 gives 604 m/ha, twice that of 004, because it is two separate pieces, and that is why the ring caveat weighs more on it.",
        "Views per tree is how many raw detections were fused into each final tree. "
        "The grid gives 9 views per point, and the measured value ranges from 9.8 to 13.8. Going over 9 means "
        "that within a single clip the network sometimes splits a tree into more than one piece, "
        "and the fusion puts it back together. 001, at 13.8, is the one that fragments the most.",
        "The other 8 stands of the farm do not appear because they have no cloud at all. Measured "
        "cell by cell: 0% coverage, except for a 100 m² strip of stand 010.",
    ]:
        c = ws.cell(ln, 1, txt)
        c.font = Font(bold=(txt == "How to read"), sz=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=8)
        if txt != "How to read":
            ws.row_dimensions[ln].height = 15 * (1 + len(txt) // 105)
        ln += 1

    wb.save(livro)
    print(f"tab '{ABA}' written to {livro.name}. Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
