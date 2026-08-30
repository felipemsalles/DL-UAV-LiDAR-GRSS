"""Trim the detection spreadsheet down to something readable at a glance.

The running text belongs in the README tab, not inside the data tabs. What this
script does:
  * rewrites the "Method summary" tab into the one that answers everything at a
    glance, with all methods in one table and the matched metric in another;
  * deletes "SegmentAnyTree" and "Test-time adaptation", whose content is in the
    summary and whose running text belongs in the README;
  * strips the paragraphs out of "13 plots", which is left with the table only.

Every count is taken at the 11.28 m radius, which closes exactly the 400 m² the field
crew measured. A 12 m radius covers 452 m² and inflates the rates by about 13%. Local
maxima and watershed are counted on both rulers by `exp_watershed_baseline.py`.

Usage: PYTHONPATH=. python scripts/xlsx_enxugar.py [workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from greenvista import config

PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"
VERDE = PatternFill("solid", fgColor="1D6B45")
CLARO = PatternFill("solid", fgColor="EDF3EF")
DESTAQUE = PatternFill("solid", fgColor="D6E9DD")
BRANCO = Font(bold=True, color="FFFFFF", sz=10)
CAMPO = 717

# name, stems within the 11.28 m radius, mean per-plot error (None = not measured),
# matched F1 on stand 001 (None = not measured), highlight
#
# The FF3D F1 values are at its own fusion radius, 1.1 m; 0.791 and 0.828 are the
# values at SegmentAnyTree's 1.5 m radius. Comparing both models at a single radius
# penalises FF3D by about 3.6 F1 points and 11 stems of count.
# Source: figs_en/table1_matched_metric.csv.
#
# Local maxima appears under two settings. The 342 comes from a single raster
# setting that was never swept; sweeping resolution and window by the same criterion
# that picks the fusion radius, the same classical algorithm finds 701. Reporting the
# 342 alone would overstate the advantage of the networks.
# Source: figs_en/table2_pooled_detection.csv.
#
# "One pass" is view g4, the centred one, filtered out of the nine; g0 is the corner
# and yields 105 and 205 instead of 388 and 655. The raw per-view detections live in
# `data/detections/`, and the two rows reconstructed this way reproduce the exact
# total, in seconds and without a GPU.
#
# All nine totals were checked against the source before the per-plot error was
# written here.
METODOS = [
    # name                                        n    MAE   F1     highlight  MAE source
    ("Local maxima, single unswept setting",    342, 28.8, None,  False),   # watershed_sweep.maximo_local_r400
    ("Watershed on the height model",           344, 28.7, None,  False),   # watershed_sweep.ws_sem_fusao_r400
    ("FF3D, one pass",                          388, 25.3, 0.685, False),   # ff3d_overlap_centroids_13plot, tile *_g4
    ("FF3D, nine views",                        549, 13.5, 0.828, False),   # ff3d_base_13parcelas, radius 1.1
    ("FF3D, nine views + AdaBN",                630, 8.2,  0.864, True),    # ff3d_adabn_13parcelas, radius 1.1
    ("SegmentAnyTree, one pass",                655, 8.3,  0.779, False),   # sat_13parcelas_instancias, baseline, *_g4
    ("SegmentAnyTree, nine views",              663, 6.3,  0.870, False),   # sat_adabn_13parcelas.base_r400
    ("Local maxima, window swept by F1",        701, 10.0, 0.792, False),   # flight_study_tres_modelos, full
    ("SegmentAnyTree, nine views + AdaBN",      737, 5.5,  0.921, True),    # sat_adabn_13parcelas.adabn_r400
]


def cab(ws, ln, cols, larguras=None):
    for j, t in enumerate(cols, 1):
        c = ws.cell(ln, j, t)
        c.fill, c.font = VERDE, BRANCO
        c.alignment = Alignment(wrap_text=True, vertical="center")
    return ln + 1


def main():
    livro = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    wb = openpyxl.load_workbook(livro)

    # ---------------------------------------------- 1) the summary, the main tab
    if "Method summary" in wb.sheetnames:
        del wb["Method summary"]
    ws = wb.create_sheet("Method summary", 1)   # right after the README
    for col, w in zip("ABCDE", [38, 12, 10, 12, 20]):
        ws.column_dimensions[col].width = w

    ws.cell(1, 1, "How many trees each method found").font = Font(bold=True, sz=14)
    ws.cell(2, 1, f"13 plots in 6 stands, {CAMPO} trees counted in the field. "
                  f"Count in the 400 m² circle, the same area the field crew measured.").font = Font(sz=11)
    ln = cab(ws, 4, ["Method", "Trees", "of 717", "Hit rate",
                     "Mean error per plot"])
    for nome, n, mae, _f1, forte in sorted(METODOS, key=lambda m: m[1]):
        fill = DESTAQUE if forte else CLARO
        for j, v in enumerate([nome, n, CAMPO, n / CAMPO, mae], 1):
            c = ws.cell(ln, j, v)
            c.fill = fill
            if j == 4:
                c.number_format = "0.0%"
            if j == 5 and v is not None:
                c.number_format = "0.0"
            if forte:
                c.font = Font(bold=True)
        ln += 1

    # the per-plot error column exists because the hit-rate column misleads: swept
    # local maxima closes 97.8% of the pooled census yet misses ten stems in every
    # plot, because the plots cancel each other out.
    ln += 1
    c = ws.cell(ln, 1, "The hit-rate column is the ratio of the pooled total, and it separates the methods badly. "
                       "Swept local maxima closes 97.8% and misses ten trees in every plot, "
                       "because the plots cancel each other out. What separates them is the per-plot error, "
                       "and after it the matched metric below.")
    c.font = Font(sz=10, italic=True)
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=5)
    ws.row_dimensions[ln].height = 42
    ln += 1

    ln += 1
    ws.cell(ln, 1, "And whether they are the right trees, in stand 001").font = Font(bold=True, sz=12)
    ln += 1
    ws.cell(ln, 1, "The only stand with a terrestrial laser stem survey. "
                   "F1 combines finding and not inventing.").font = Font(sz=10, italic=True)
    ln += 1
    ln = cab(ws, ln, ["Method", "F1"])
    # sort by F1, not by the list order, which is the count order. The two orders do
    # not coincide, and that is precisely the finding: swept local maxima is nearly the
    # best on count yet sits in the middle of the table on quality.
    for nome, _n, _mae, f1, forte in sorted(METODOS, key=lambda m: m[3] or 0):
        if f1 is None:
            continue
        fill = DESTAQUE if forte else CLARO
        for j, v in enumerate([nome, f1], 1):
            c = ws.cell(ln, j, v)
            c.fill = fill
            if j == 2:
                c.number_format = "0.000"
            if forte:
                c.font = Font(bold=True)
        ln += 1

    ln += 1
    c = ws.cell(ln, 1, "Each network is at its own fusion radius, 1.1 m for FF3D and 1.5 m for "
                       "SegmentAnyTree. Comparing the two at a single radius takes 3.6 F1 points off FF3D. "
                       "A 12 m circle covers 452 m² and inflates the rates by about 13%; "
                       "everything here is at the 11.28 m radius, "
                       "which closes the 400 m² the field crew measured.")
    c.font = Font(sz=10, italic=True)
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=5)
    ws.row_dimensions[ln].height = 56

    # ---------------------------------------------- 2) tabs that go away
    for morta in ("SegmentAnyTree", "Test-time adaptation"):
        if morta in wb.sheetnames:
            del wb[morta]

    # ---------------------------------------------- 3) strip prose out of "13 plots"
    if "13 plots" in wb.sheetnames:
        w2 = wb["13 plots"]
        # the paragraphs were written into merged cells, and MergedCell does not
        # accept assignment; unmerge before clearing.
        for rng in list(w2.merged_cells.ranges):
            w2.unmerge_cells(str(rng))
        for r in list(w2.iter_rows()):
            prim = r[0].value
            if isinstance(prim, str) and len(prim) > 70 and all(c.value is None for c in r[1:]):
                for c in r:
                    c.value = None
        w2.cell(3, 1, "Count in the 400 m² circle. Each model at its own fusion radius.")
        w2.cell(3, 1).font = Font(sz=10, italic=True)

    # nothing is appended to the README here: appending with `rd.cell(rd.max_row + 2, ...)`
    # adds to the end of the tab on every execution and duplicates the text over repeated
    # runs. The README is written from scratch by `xlsx_readme.py`, which owns that tab.

    wb.save(livro)
    print(f"workbook trimmed: {len(wb.sheetnames)} tabs -> {wb.sheetnames}")


if __name__ == "__main__":
    main()
