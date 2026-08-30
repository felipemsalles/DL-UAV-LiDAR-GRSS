"""Rebuild the "Per stand" and "Per plot" tabs with every method, at the 11.28 m radius.

The two tabs become the single source per plot and per stand again, and the
"13 plots" tab, which duplicated the same count with a different set of methods, is
deleted.

The detail tabs carry three methods: the classical one as a reference, plus the best of
each network family; the full ladder of eight stays in the summary.

Everything at the 11.28 m circle, which closes the 400 m² measured in the field.

Usage: PYTHONPATH=. python scripts/xlsx_por_parcela_talhao.py [workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from greenvista import config

PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"
VERDE = PatternFill("solid", fgColor="1D6B45")
CLARO = PatternFill("solid", fgColor="EDF3EF")
BRANCO = Font(bold=True, color="FFFFFF", sz=10)
NOME_TALHAO = {1: "001", 2: "002", 4: "004 (young)", 5: "005", 6: "006", 7: "007"}


def dados():
    ws_cl = pd.read_csv(config.OUT_DIR / "watershed_sweep.csv")
    ws_cl["parcela_id"] = ws_cl.talhao.astype(str) + "_" + ws_cl.parcela.astype(str)
    ws_cl = ws_cl.set_index("parcela_id")
    ff = pd.read_csv(config.OUT_DIR / "ff3d_adabn_13parcelas.csv")
    ff = ff[ff.raio == 1.1].set_index("parcela")
    sat = pd.read_csv(config.OUT_DIR / "sat_adabn_13parcelas.csv").set_index("parcela")
    linhas = []
    for pid in ff.index:
        linhas.append(dict(
            parcela=pid, talhao=int(pid.split("_")[0]), campo=int(ff.loc[pid, "campo"]),
            maximo_local=int(ws_cl.loc[pid, "maximo_local_r400"]),
            ff3d=int(ff.loc[pid, "n_r400"]),
            sat=int(sat.loc[pid, "adabn_r400"])))
    d = pd.DataFrame(linhas)
    return d.sort_values(["talhao", "parcela"], key=lambda s: s.map(
        lambda v: (int(str(v).split("_")[1]) if "_" in str(v) else v)) if s.name == "parcela" else s)


def cab(ws, ln, fixas, metodos, extra=None):
    """Two-storey header.

    On top the method name, spanning two columns; underneath "trees" and
    "hit rate". The fixed columns (plot, stand, field) and the remarks column
    are merged vertically across both storeys, so that "hit rate" never stands alone
    without an indication of which method it belongs to.
    """
    meio = Alignment(horizontal="center", vertical="center", wrap_text=True)
    j = 1
    for t in fixas:
        ws.merge_cells(start_row=ln, start_column=j, end_row=ln + 1, end_column=j)
        c = ws.cell(ln, j, t); c.fill, c.font, c.alignment = VERDE, BRANCO, meio
        ws.cell(ln + 1, j).fill = VERDE
        j += 1
    for t in metodos:
        ws.merge_cells(start_row=ln, start_column=j, end_row=ln, end_column=j + 1)
        c = ws.cell(ln, j, t); c.fill, c.font, c.alignment = VERDE, BRANCO, meio
        ws.cell(ln, j + 1).fill = VERDE
        for k, sub in enumerate(("trees", "hit rate")):
            c = ws.cell(ln + 1, j + k, sub)
            c.fill, c.font, c.alignment = VERDE, BRANCO, meio
        j += 2
    if extra:
        ws.merge_cells(start_row=ln, start_column=j, end_row=ln + 1, end_column=j)
        c = ws.cell(ln, j, extra); c.fill, c.font, c.alignment = VERDE, BRANCO, meio
        ws.cell(ln + 1, j).fill = VERDE
    ws.row_dimensions[ln].height = 30
    return ln + 2


def corpo(ws, ln, vals, negrito=False):
    for j, v in enumerate(vals, 1):
        c = ws.cell(ln, j, v)
        c.fill = CLARO
        if isinstance(v, float):
            c.number_format = "0.0%"
        if negrito:
            c.font = Font(bold=True)
    return ln + 1


# the label says "unswept" because the summary carries two local-maxima
# settings, the historical one and the F1-swept one, which give 342 against 701 stems;
# without the qualifier there is no telling which of the two is in the column.
COLS = ["Local maxima (unswept)", "FF3D + AdaBN", "SegmentAnyTree + AdaBN"]
CHAVES = ["maximo_local", "ff3d", "sat"]


def escreve(ws, titulo, subtitulo, primeira, grupos, obs=None):
    ws.cell(1, 1, titulo).font = Font(bold=True, sz=14)
    ws.cell(2, 1, subtitulo).font = Font(sz=10, italic=True)
    ln = cab(ws, 4, list(primeira), COLS, "Remarks" if obs is not None else None)
    for chave, linha in grupos:
        vals = list(chave)
        for k in CHAVES:
            vals += [linha[k], linha[k] / linha["campo"]]
        if obs is not None:
            vals.append(obs.get(chave[0], ""))
        ln = corpo(ws, ln, vals)
    return ln


def main():
    livro = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    d = dados()
    wb = openpyxl.load_workbook(livro)

    # keep the remarks that were already there
    obs = {}
    if "Per plot" in wb.sheetnames:
        antiga = wb["Per plot"]
        for r in antiga.iter_rows(min_row=6, values_only=True):
            if r and r[0] and r[-1]:
                obs[str(r[0])] = str(r[-1])

    for nome in ("Per plot", "Per stand", "13 plots"):
        if nome in wb.sheetnames:
            del wb[nome]

    # ------------------------------------------------------------------- plot
    ws = wb.create_sheet("Per plot", 2)
    linhas = [((r.parcela, NOME_TALHAO[r.talhao], r.campo), r._asdict())
              for r in d.itertuples()]
    ln = escreve(ws, "Trees found per plot",
                 "Plots of 400 m². Count in the 11.28 m circle, which closes that same area.",
                 ("Plot", "Stand", "Field"), linhas, obs)
    tot = {k: int(d[k].sum()) for k in CHAVES}; tot["campo"] = int(d.campo.sum())
    corpo(ws, ln, ["Total", "", tot["campo"]] +
          [x for k in CHAVES for x in (tot[k], tot[k] / tot["campo"])] + [""], negrito=True)
    for col, w in zip("ABCDEFGHIJ", [9, 12, 8, 12, 9, 13, 9, 20, 9, 40]):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------------------------ stand
    ws = wb.create_sheet("Per stand", 3)
    g = d.groupby("talhao").agg(parcelas=("parcela", "size"), campo=("campo", "sum"),
                                **{k: (k, "sum") for k in CHAVES}).reset_index()
    linhas = [((NOME_TALHAO[int(r.talhao)], int(r.parcelas), int(r.campo)), r._asdict())
              for r in g.itertuples()]
    ln = escreve(ws, "Trees found per stand",
                 "Sum of the plots of each stand, and not of the whole stand. "
                 "Outside the plots there is no field count to compare against.",
                 ("Stand", "Plots", "Field"), linhas)
    corpo(ws, ln, ["Total", int(g.parcelas.sum()), tot["campo"]] +
          [x for k in CHAVES for x in (tot[k], tot[k] / tot["campo"])], negrito=True)
    for col, w in zip("ABCDEFGHI", [13, 10, 8, 12, 9, 13, 9, 20, 9]):
        ws.column_dimensions[col].width = w

    wb.save(livro)
    print(f"tabs rebuilt. workbook now has {len(wb.sheetnames)}: {wb.sheetnames}")


if __name__ == "__main__":
    main()
