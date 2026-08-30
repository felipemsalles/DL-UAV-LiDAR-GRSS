"""Write the 13-plot tab into the detection workbook.

Carries both models, aggregated and with AdaBN, over the 13 plots of the 7 stands, with
9 views in all of them. Outside stand 001 there is no stem map, so all that is available
here is the count against the field census.

Usage: PYTHONPATH=. python scripts/xlsx_aba_13parcelas.py [workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from greenvista import config

ABA = "13 plots"
PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"
VERDE = PatternFill("solid", fgColor="1D6B45")
CLARO = PatternFill("solid", fgColor="EDF3EF")
BRANCO = Font(bold=True, color="FFFFFF", sz=10)


def cab(ws, ln, cols):
    for j, t in enumerate(cols, 1):
        c = ws.cell(ln, j, t)
        c.fill, c.font = VERDE, BRANCO
        c.alignment = Alignment(wrap_text=True, vertical="center")


def linha(ws, ln, vals, fmts=None):
    for j, v in enumerate(vals, 1):
        f = fmts[j - 1] if fmts else None
        if f and f.endswith("%"):
            v = float(v)
        c = ws.cell(ln, j, v)
        c.fill = CLARO
        if f:
            c.number_format = f


def tit(ws, ln, txt, sz=11):
    ws.cell(ln, 1, txt).font = Font(bold=True, sz=sz)
    return ln + 1


def main():
    livro = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    ff = pd.read_csv(config.OUT_DIR / "ff3d_adabn_13parcelas.csv")
    sat = pd.read_csv(config.OUT_DIR / "sat_adabn_13parcelas.csv")
    ffb = pd.read_csv(config.OUT_DIR / "ff3d_base_13parcelas.csv")
    ffb = ffb[ffb.raio == 1.1].set_index("parcela")

    wb = openpyxl.load_workbook(livro)
    if ABA in wb.sheetnames:
        del wb[ABA]
    ws = wb.create_sheet(ABA)
    for col, w in zip("ABCDEFGH", [10, 9, 15, 13, 15, 13, 15, 13]):
        ws.column_dimensions[col].width = w

    ln = tit(ws, 1, "The 13 plots, the two models, with and without adaptation", sz=14)
    ln = tit(ws, ln, "7 stands, 717 census trees, 9 views of 32 m in every plot. "
                     "Outside stand 001 there is no stem map, so all there is here is count, "
                     "which says how many trees and not which ones.")
    ln = tit(ws, ln, "Counting radius 11.28 m, which closes the 400 m² the field crew measured. Each "
                     "model at its own fusion radius: 1.1 m for FF3D and 1.5 m for SegmentAnyTree. "
                     "Using a single radius for both penalises FF3D by about 11 points.")
    ln += 1

    cab(ws, ln, ["Plot", "Field",
                 "FF3D aggregated", "hit rate",
                 "FF3D + AdaBN", "hit rate",
                 "SegmentAnyTree + AdaBN", "hit rate"])
    ln += 1
    f11 = ff[ff.raio == 1.1].set_index("parcela")
    s = sat.set_index("parcela")
    ordem = sorted(f11.index, key=lambda p: (int(p.split("_")[0]), int(p.split("_")[1])))
    for p in ordem:
        campo = int(f11.loc[p, "campo"])
        a = int(f11.loc[p, "n_r400"])
        b = int(ffb.loc[p, "n_r400"])
        sa = int(s.loc[p, "adabn_r400"])
        linha(ws, ln, [p, campo, b, b / campo, a, a / campo, sa, sa / campo],
              [None, None, None, "0.0%", None, "0.0%", None, "0.0%"])
        ln += 1
    C = int(f11.campo.sum())
    linha(ws, ln, ["TOTAL", C, int(ffb.n_r400.sum()), ffb.n_r400.sum() / C,
                   int(f11.n_r400.sum()), f11.n_r400.sum() / C,
                   int(s.adabn_r400.sum()), s.adabn_r400.sum() / C],
          [None, None, None, "0.0%", None, "0.0%", None, "0.0%"])
    for j in range(1, 9):
        ws.cell(ln, j).font = Font(bold=True)
    ln += 2

    ln = tit(ws, ln, "Summary, the 13 plots pooled")
    cab(ws, ln, ["Model", "Fusion radius", "Aggregated", "hit rate", "Aggregated + AdaBN", "hit rate"])
    ln += 1
    for nome, raio, agr, ada in [
            ("FF3D", 1.5, 468, int(ff[ff.raio == 1.5].n_r400.sum())),
            ("FF3D", 1.1, 549, int(f11.n_r400.sum())),
            ("SegmentAnyTree", 1.5, int(s.base_r400.sum()), int(s.adabn_r400.sum()))]:
        linha(ws, ln, [nome, raio, agr, agr / C, ada, ada / C],
              [None, "0.0", None, "0.0%", None, "0.0%"])
        ln += 1
    ln += 1

    for t in ["How to read",
              "AdaBN helps both almost equally, 11.3 points for FF3D and 10.3 for SegmentAnyTree, "
              "each at its own radius. The effect is robust, not a quirk of one model.",
              "The fusion radius is worth almost as much as adaptation, and only for FF3D. Swapping 1.5 "
              "for 1.1 gives it 11 points and gives SegmentAnyTree zero.",
              "SegmentAnyTree stays ahead, by less than it seemed before: 102.8% against "
              "87.9% at each one's best setting.",
              "Plot 7_14 is the one that cannot be defended. A census of 33, the smallest of all, and both "
              "models go over 100%. With no stem map there, coppice regrowth cannot be told apart "
              "from model error.",
              "Source: manual_match/ff3d_adabn_13parcelas.csv and sat_adabn_13parcelas.csv"]:
        c = ws.cell(ln, 1, t)
        c.font = Font(bold=(t == "How to read"), sz=11)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=ln, start_column=1, end_row=ln, end_column=8)
        ln += 1

    ws.freeze_panes = "A2"
    wb.save(livro)
    print(f"tab '{ABA}' written to {livro}")


if __name__ == "__main__":
    main()
