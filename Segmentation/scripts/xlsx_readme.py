"""Rewrite the README tab of the detection workbook from scratch.

The rewrite is idempotent: the tab is deleted and recreated, so running it twice
duplicates no content.

Usage: PYTHONPATH=. python scripts/xlsx_readme.py [workbook.xlsx]
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from greenvista import config

PADRAO = config.REPO.parent / "deteccao-arvores-por-talhao-e-parcela.xlsx"

TITULO = "Automatic tree detection from drone LiDAR"
SUB = "Eucalyptus, Fazenda São Manuel. Undergraduate research, ICMC USP, Bruno Araujo."

# (is it a section heading?, text)
LINHAS = [
    (True, "What is in here"),
    (False, "How many trees each automatic method found, compared with the count made "
            "in the field. There are 13 plots of 400 m² in 6 stands, with 717 measured trees."),

    (True, "The methods"),
    (False, "Local maxima and watershed look for crown tops in a height model. They are the "
            "classical methods, and it is against them that the networks have to justify themselves."),
    (False, "The local maxima setting changes everything, and that is why it appears twice in the "
            "table. Under a single setting, chosen without a sweep, it finds 342 trees. Sweeping "
            "the raster resolution and the window size and keeping the best, the same algorithm "
            "finds 701. Reporting only the first would make the networks look far better than they are."),
    (False, "FF3D and SegmentAnyTree are networks that segment the point cloud tree by tree. "
            "Both were trained on natural forest, and neither has seen this plantation."),

    (True, "Three things that change the result a lot"),
    (False, "Nine views instead of one. Each plot is clipped nine times with an offset, and the "
            "detections are fused. A tree on the edge of one clip falls in the middle of another."),
    (False, "Test-time adaptation. The network recalibrates on the cloud it is seeing, with no "
            "labels needed. It is worth about 11 points, and it holds for both networks."),
    (False, "Fusion radius. When the nine views are merged, nearby detections collapse into one. "
            "The good radius is not the same for the two networks, 1.1 m for FF3D and 1.5 m for "
            "SegmentAnyTree. Comparing the two at a single radius penalises FF3D by about 11 points."),

    (True, "The caveat that matters most"),
    (False, "Count says how many trees, not which ones. If the computer lets 5 slip and invents "
            "another 5, the total comes out right while being wrong 10 times."),
    (False, "This is not a hypothesis. Swept local maxima closes 97.8% of the pooled census and "
            "misses ten trees in every plot, against 5.5 for SegmentAnyTree, because the plots "
            "cancel each other out. That is why the table carries the per-plot error column "
            "next to the hit-rate one."),
    (False, "Only stand 001 can be checked, because only it has a terrestrial laser stem survey, "
            "with 892 trees. There the comparison is tree by tree, and that is why the summary "
            "tab carries a second block for it alone."),

    (True, "About going over 100%"),
    (False, "SegmentAnyTree with adaptation goes past the census in several plots. In stand 001, "
            "where it can be checked, those extra detections are real trees, at 92.8% "
            "precision, which means the field census is the one falling short of what the cloud "
            "shows. Outside 001 the same cannot be claimed."),

    (True, "The whole-stand number"),
    (False, "It exists, and it is in the \"Whole stand\" tab. The six stands with a cloud were "
            "swept in full, 9.87 hectares, with the same nine-view scheme used on the "
            "plots."),
    (False, "It is not verified, and that changes how it is read. The inventory has 717 trees and "
            "all of them fall inside the 13 plots, none outside. So the stand total is a model "
            "measurement, not a checked measurement. What can be checked is the part falling in the plots."),
    (False, "The edge deserves suspicion. The clouds were clipped exactly at the stand "
            "boundary, with no metre of margin, measured point by point. A tree in the outer row has "
            "its crown truncated for lack of data. The tab separates how much of the count comes from that ring."),
    (False, "The other eight stands of the farm still have no number, and there it really is "
            "missing data. LiDAR coverage measured cell by cell: zero, except for a 100 m² strip of "
            "stand 010."),

    (True, "About the counting radius"),
    (False, "Everything here is counted in an 11.28 m circle, which closes the same 400 m² the "
            "field crew measured. A 12 m circle covers 452 m² and inflates every rate by "
            "about 10 points."),
]


def main():
    livro = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    wb = openpyxl.load_workbook(livro)
    if "README" in wb.sheetnames:
        del wb["README"]
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 118

    ws.cell(1, 1, TITULO).font = Font(bold=True, sz=14)
    ws.cell(2, 1, SUB).font = Font(sz=11)
    ln = 4
    for secao, txt in LINHAS:
        c = ws.cell(ln, 1, txt)
        if secao:
            ln += 0 if ln == 4 else 0
            c.font = Font(bold=True, sz=12)
        else:
            c.font = Font(sz=11)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[ln].height = 15 * (1 + len(txt) // 110)
        ln += 1
        if secao:
            continue
    wb.save(livro)
    print(f"README rewritten, {ln - 4} lines")


if __name__ == "__main__":
    main()
